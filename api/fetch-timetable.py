"""
Vercel Serverless Function / GitHub Actions CLI: fetch-timetable

Fetches the latest matching Gmail message and routes it by subject keyword:
  - subject contains "seating"  -> parses the attached seating-plan PDF into
    dbfolder/seating-plan.json
  - subject contains "schedule" -> parses the attached Final Exam Schedule
    .xlsx workbook (date/time-slot matrix layout) into
    dbfolder/exam-schedule-<school>.json
  - subject contains "showup"   -> parses the attached Show Up Schedule .xlsx
    workbook (plain one-row-per-section table, different layout) into
    dbfolder/showup-schedule-<school>.json. The school edits this data
    directly in a linked Google Sheet rather than resending email, so this
    module also extracts that sheet's link from the email body once (see
    maybe_bootstrap_showup_sheet_source) and `--poll-showup` (run via
    .github/workflows/poll-showup-sheet.yml on a 5-min cron) re-fetches that
    live sheet directly, independent of any new email.

Required environment variables:
  GMAIL_USER  – Gmail address
  GMAIL_PASS  – Gmail App Password (not your regular password)
  GH_TOKEN    – GitHub PAT with repo contents write access (Vercel handler only;
                the GitHub Actions CLI commits with its own built-in token)

Optional environment variables:
  GH_REPO     – "owner/repo" (defaults to VERCEL_GIT_REPO_OWNER/SLUG)
  GH_BRANCH   – target branch (default: main)
"""

from __future__ import annotations

import base64
import email
import imaplib
import json
import os
import re
import shutil
import ssl
import io
import subprocess
import tempfile
import pypdf
from datetime import datetime, timezone
from email.header import decode_header
from http.server import BaseHTTPRequestHandler
from typing import Any
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

# ── Constants ────────────────────────────────────────────────────────────────

GMAIL_IMAP_HOST = "imap.gmail.com"
GMAIL_IMAP_PORT = 993
# Subject keyword -> (kind label, JSON output path). Matched case-insensitively
# against the email subject; kept broad so "Seating Plan", "seatingplan",
# "Schedule of Final Examination", etc. all match.
SUBJECT_ROUTES: dict[str, tuple[str, str]] = {
    "seating": ("seating", "dbfolder/seating-plan.json"),
    "schedule": ("exam_schedule", "dbfolder/exam-schedule-computing.json"),
    "showup": ("showup_schedule", "dbfolder/showup-schedule-computing.json"),
}

# For the multi-school xlsx routes, the per-school output path is
# dbfolder/{prefix}-{school}.json — this maps kind -> that filename prefix.
SCHEDULE_FILE_PREFIX: dict[str, str] = {
    "exam_schedule": "exam-schedule",
    "showup_schedule": "showup-schedule",
}

# Column header aliases → canonical JSON field names
FIELD_ALIASES: dict[str, str] = {
    "name": "name",
    "student name": "name",
    "student": "name",
    "nuid": "nuid",
    "nu id": "nuid",
    "nu-id": "nuid",
    "id": "nuid",
    "roll": "nuid",
    "roll no": "nuid",
    "roll number": "nuid",
    "paper": "paper",
    "subject": "paper",
    "course": "paper",
    "exam": "paper",
    "paper name": "paper",
    "time": "time",
    "exam time": "time",
    "slot": "time",
    "class": "class",
    "room": "class",
    "hall": "class",
    "venue": "class",
    "class room": "class",
    "classroom": "class",
    "seat": "seat",
    "seat no": "seat",
    "seat number": "seat",
    "seatno": "seat",
    "position": "seat",
}

KV_ALIASES: dict[str, str] = {
    "name": "name",
    "student name": "name",
    "nuid": "nuid",
    "nu id": "nuid",
    "nu-id": "nuid",
    "roll": "nuid",
    "paper": "paper",
    "subject": "paper",
    "course": "paper",
    "time": "time",
    "class": "class",
    "room": "class",
    "hall": "class",
    "seat": "seat",
    "seat no": "seat",
    "seat number": "seat",
}


# ── Response helpers ───────────────────────────────────────────────────────────

def json_response(handler: BaseHTTPRequestHandler, status: int, payload: dict) -> None:
    body = json.dumps(payload, indent=2).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


# ── Gmail IMAP & PDF Extraction ─────────────────────────────────────────────────

def decode_mime_header(value: str) -> str:
    """Decode RFC 2047 encoded email headers into a plain string."""
    parts = decode_header(value or "")
    decoded: list[str] = []
    for fragment, charset in parts:
        if isinstance(fragment, bytes):
            decoded.append(fragment.decode(charset or "utf-8", errors="replace"))
        else:
            decoded.append(fragment)
    return "".join(decoded).strip()


def extract_pdf_text(payload: bytes) -> str:
    """Extract text from a PDF, preserving column layout.

    FAST seating sheets are two-column positional tables, so column layout must
    be kept. Prefer `pdftotext -layout` (poppler); fall back to pypdf's layout
    mode if the binary isn't present (e.g. on Vercel).
    """
    if shutil.which("pdftotext"):
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
                tmp.write(payload)
                tmp_path = tmp.name
            result = subprocess.run(
                ["pdftotext", "-layout", tmp_path, "-"],
                capture_output=True, timeout=60,
            )
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.decode("utf-8", errors="replace")
        except Exception:
            pass
        finally:
            if tmp_path and os.path.exists(tmp_path):
                os.unlink(tmp_path)

    text_content = []
    try:
        reader = pypdf.PdfReader(io.BytesIO(payload))
        for page in reader.pages:
            try:
                page_text = page.extract_text(extraction_mode="layout")
            except TypeError:
                page_text = page.extract_text()
            if page_text:
                text_content.append(page_text)
        return "\n".join(text_content)
    except Exception as e:
        raise RuntimeError(f"Failed to process and parse PDF attachment data: {str(e)}")


def extract_plain_text(msg: email.message.Message) -> str:
    """
    Walk the MIME tree and isolate PDF attachment blocks. If a PDF file is present,
    it maps out text from it. Otherwise, falls back to raw plain/html fallback trees.
    """
    # High Priority Pass: Intercept PDF attachment configurations
    for part in msg.walk():
        content_type = part.get_content_type()
        filename = part.get_filename() or ""
        
        if content_type == "application/pdf" or filename.lower().endswith(".pdf"):
            payload = part.get_payload(decode=True)
            if payload:
                pdf_text = extract_pdf_text(payload)
                if pdf_text.strip():
                    return pdf_text

    # Default Low Priority Pass: Parse standard text body components
    if msg.is_multipart():
        plain_parts: list[str] = []
        html_parts: list[str] = []
        for part in msg.walk():
            content_type = part.get_content_type()
            disposition = str(part.get("Content-Disposition", ""))
            if "attachment" in disposition.lower():
                continue
            payload = part.get_payload(decode=True)
            if payload is None:
                continue
            charset = part.get_content_charset() or "utf-8"
            text = payload.decode(charset, errors="replace")
            if content_type == "text/plain":
                plain_parts.append(text)
            elif content_type == "text/html":
                html_parts.append(text)
        if plain_parts:
            return "\n".join(plain_parts)
        if html_parts:
            return strip_html("\n".join(html_parts))
        return ""
        
    payload = msg.get_payload(decode=True)
    if payload is None:
        return str(msg.get_payload())
    charset = msg.get_content_charset() or "utf-8"
    text = payload.decode(charset, errors="replace")
    if msg.get_content_type() == "text/html":
        return strip_html(text)
    return text


def strip_html(html: str) -> str:
    """Minimal HTML → text conversion for seating-plan table emails."""
    text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", html)
    # Preserve link targets inline (e.g. "View Sheet (https://docs.google.com/...)")
    # before the generic tag-stripper below discards them - a "click here"-style
    # link would otherwise silently lose the actual URL.
    text = re.sub(r'(?is)<a\s+[^>]*href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', r"\2 (\1)", text)
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</tr>", "\n", text)
    text = re.sub(r"(?i)</t[dh]>", "\t", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _build_subject_or_criteria(keywords: list[str]) -> str:
    """Build an IMAP SEARCH criteria string matching ANY of the given subject
    keywords, e.g. ["seating", "schedule"] -> '(OR (SUBJECT "seating") (SUBJECT "schedule"))'."""
    chain = f'SUBJECT "{keywords[0]}"'
    for kw in keywords[1:]:
        chain = f'OR ({chain}) (SUBJECT "{kw}")'
    return f"({chain})"


def fetch_latest_matching_email() -> tuple[str, str, str, bytes | None, str]:
    """
    Connect to Gmail IMAP, find the newest message whose subject matches one of
    SUBJECT_ROUTES (preferring unread), and return (kind, subject, plain_body,
    attachment_bytes, attachment_filename). `kind` is "seating" or
    "exam_schedule" per SUBJECT_ROUTES; attachment_filename is "" when not
    applicable (e.g. the seating PDF route doesn't need it).
    Marks the message as \\Seen after a successful read.
    """
    user = os.environ.get("GMAIL_USER", "").strip()
    password = os.environ.get("GMAIL_PASS", "").strip()
    if not user or not password:
        raise RuntimeError("GMAIL_USER and GMAIL_PASS environment variables are required")

    context = ssl.create_default_context()
    mail = imaplib.IMAP4_SSL(GMAIL_IMAP_HOST, GMAIL_IMAP_PORT, ssl_context=context)
    try:
        mail.login(user, password)
        mail.select("INBOX")

        subject_criteria = _build_subject_or_criteria(list(SUBJECT_ROUTES.keys()))

        # Prefer the newest UNREAD matching email; if none is found (e.g. it was
        # already opened), fall back to the newest matching email regardless of
        # read state so an opened email still gets processed.
        ids: list[bytes] = []
        for criteria in (
            ("UNSEEN", subject_criteria),
            (subject_criteria,),
        ):
            status, data = mail.search(None, *criteria)
            if status != "OK":
                raise RuntimeError(f"IMAP search failed: {status}")
            found = data[0].split() if data and data[0] else []
            if found:
                ids = found
                break

        if not ids:
            raise RuntimeError('No emails found with "seating", "schedule", or "showup" in the subject')

        latest_id = ids[-1]
        status, fetched = mail.fetch(latest_id, "(RFC822)")
        if status != "OK" or not fetched or not fetched[0]:
            raise RuntimeError("Failed to fetch email payload")

        raw_bytes = fetched[0][1]
        msg = email.message_from_bytes(raw_bytes)
        subject = decode_mime_header(msg.get("Subject", ""))
        subject_lower = subject.lower()

        kind = None
        for keyword, (label, _path) in SUBJECT_ROUTES.items():
            if keyword in subject_lower:
                kind = label
                break
        if kind is None:
            raise RuntimeError(f"Matched email subject {subject!r} did not match any known route")

        # Always extract body text - for "seating" this is (mainly) the PDF's
        # own text; for xlsx routes there's no PDF so this naturally falls
        # through to the mail's own plain/html body, which is what lets us
        # scan for a linked Google Sheet URL below (see maybe_bootstrap_showup_sheet_source).
        body = extract_plain_text(msg)
        attachment_filename = ""
        if kind == "seating":
            attachment = find_pdf_attachment(msg)
        else:
            xlsx_found = find_xlsx_attachment(msg)
            if xlsx_found:
                attachment, attachment_filename = xlsx_found
            else:
                attachment = None

        mail.store(latest_id, "+FLAGS", "\\Seen")
        return kind, subject, body, attachment, attachment_filename
    finally:
        try:
            mail.logout()
        except Exception:
            pass


# ── Email body → JSON parsing ──────────────────────────────────────────────────

def normalize_header(label: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (label or "").lower()).strip()


def detect_delimiter(line: str) -> str | None:
    if "\t" in line:
        return "\t"
    if "|" in line:
        return "|"
    if "," in line and line.count(",") >= 3:
        return ","
    if ";" in line and line.count(";") >= 3:
        return ";"
    return None


def split_row(line: str, delimiter: str) -> list[str]:
    if delimiter == ",":
        return [cell.strip() for cell in line.split(",")]
    return [cell.strip() for cell in line.split(delimiter)]


def map_header_indices(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, header in enumerate(headers):
        canonical = FIELD_ALIASES.get(normalize_header(header))
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
    return mapping


def row_to_entry(cells: list[str], col_map: dict[str, int]) -> dict[str, str] | None:
    entry: dict[str, str] = {}
    for field, idx in col_map.items():
        if idx < len(cells):
            value = cells[idx].strip()
            if value:
                entry[field] = value
    if entry.get("name") or entry.get("nuid"):
        return entry
    return None


def parse_tabular_body(text: str) -> list[dict[str, str]]:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if not lines:
        return []

    header_idx = -1
    delimiter: str | None = None
    col_map: dict[str, int] = {}

    for i, line in enumerate(lines):
        delim = detect_delimiter(line)
        if not delim:
            continue
        cells = split_row(line, delim)
        candidate = map_header_indices(cells)
        if ("name" in candidate or "nuid" in candidate) and (
            "paper" in candidate or "class" in candidate or "seat" in candidate
        ):
            header_idx = i
            delimiter = delim
            col_map = candidate
            break

    if header_idx < 0 or not delimiter:
        return []

    entries: list[dict[str, str]] = []
    for line in lines[header_idx + 1 :]:
        if re.match(r"^[-_=]{3,}$", line):
            continue
        cells = split_row(line, delimiter)
        if len(cells) < 2:
            continue
        entry = row_to_entry(cells, col_map)
        if entry:
            entries.append(entry)
    return entries


def parse_key_value_body(text: str) -> list[dict[str, str]]:
    blocks = re.split(r"\n\s*\n", text.strip())
    entries: list[dict[str, str]] = []

    for block in blocks:
        entry: dict[str, str] = {}
        for line in block.splitlines():
            match = re.match(r"^([^:]+):\s*(.+)$", line.strip())
            if not match:
                continue
            key = KV_ALIASES.get(normalize_header(match.group(1)))
            value = match.group(2).strip()
            if key and value:
                entry[key] = value
        if entry.get("name") or entry.get("nuid"):
            entries.append(entry)

    return entries


def normalize_seat(seat: str) -> str:
    """Normalize seat strings cleanly into standard row-col signatures."""
    raw = (seat or "").upper().strip()
    if not raw:
        return ""

    m = re.search(r"C\s*[-:]?\s*(\d+)\s*R\s*[-:]?\s*(\d+)", raw)
    if m:
        return f"C{m.group(1)}R{m.group(2)}"

    m = re.search(r"COL(?:UMN)?\s*(\d+).{0,6}ROW\s*(\d+)", raw)
    if m:
        return f"C{m.group(1)}R{m.group(2)}"

    m = re.search(r"(\d+)\s*C.*?(\d+)\s*R", raw)
    if m:
        return f"C{m.group(1)}R{m.group(2)}"

    return raw


def normalize_entry(entry: dict[str, str]) -> dict[str, str]:
    normalized = {
        "name": entry.get("name", "").strip(),
        "nuid": entry.get("nuid", "").strip().upper(),
        "paper": entry.get("paper", "").strip(),
        "time": entry.get("time", "").strip(),
        "class": entry.get("class", "").strip(),
        "seat": normalize_seat(entry.get("seat", "")),
    }
    return {k: v for k, v in normalized.items() if v}


# ── FAST exam seating-sheet parser (two-column positional layout) ──────────────
# Handles the jsPDF "SEATING PLAN" sheets: shared Date/Slot/Venue headers, a
# course/paper line, and two side-by-side "S# Roll No. Student Name Seat" tables.

# Seat may be a plain number (e.g. "21") or a column-row code (e.g. "C1R2").
FAST_REC = re.compile(r"^\s*(\d{1,3})\s+(\d{2}[A-Za-z]-\d{4})\s+(.+?)\s+([A-Z]\d+[A-Z]\d+|\d{1,3})\s*$")
FAST_PAPER = re.compile(r"^\s*([A-Z]{2,3}\d{3,4}\s*,\s*[A-Z0-9\-/]+)\s*-?\s*(.*)$")


def _clean_paper(code: str, rest: str) -> str:
    rest = re.sub(r"\s+", " ", rest or "").rstrip("- ").strip()
    code = re.sub(r"\s+", "", code)
    return f"{code} - {rest}" if rest else code


def find_pdf_attachment(msg: email.message.Message) -> bytes | None:
    """Return the raw bytes of the first PDF attachment, or None."""
    for part in msg.walk():
        content_type = part.get_content_type()
        filename = part.get_filename() or ""
        if content_type == "application/pdf" or filename.lower().endswith(".pdf"):
            payload = part.get_payload(decode=True)
            if payload:
                return payload
    return None


XLSX_CONTENT_TYPES = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
)


def find_xlsx_attachment(msg: email.message.Message) -> tuple[bytes, str] | None:
    """Return (raw_bytes, filename) of the first .xlsx/.xlsm attachment, or None."""
    for part in msg.walk():
        content_type = part.get_content_type()
        filename = part.get_filename() or ""
        if filename.lower().endswith((".xlsx", ".xlsm")) or content_type in XLSX_CONTENT_TYPES:
            payload = part.get_payload(decode=True)
            if payload:
                return payload, decode_mime_header(filename) or "exam-schedule.xlsx"
    return None


def _cluster_rows(words: list[dict], tol: float = 3.0) -> list[list[dict]]:
    """Group pdfplumber words into visual rows by their vertical position."""
    rows: list[list[dict]] = []
    for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
        for row in rows:
            if abs(row[0]["top"] - w["top"]) <= tol:
                row.append(w)
                break
        else:
            rows.append([w])
    for row in rows:
        row.sort(key=lambda w: w["x0"])
    rows.sort(key=lambda row: min(w["top"] for w in row))
    return rows


def parse_pdf_coordinates(payload: bytes) -> tuple[list[dict[str, str]], str]:
    """Coordinate-based parse of a FAST seating PDF using pdfplumber.

    Grouping words by their y-position keeps each student's seat on the same row
    even when a multi-line section header would otherwise drift the seat column
    (which breaks text-only extraction). Each page is one venue; within a page
    the two columns are read left-then-right so a section header at the bottom of
    the left column correctly applies to the right column.
    """
    import pdfplumber  # local import: text fallback still works if unavailable

    students: list[dict[str, str]] = []
    date = ""
    with pdfplumber.open(io.BytesIO(payload)) as pdf:
        for page in pdf.pages:
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            if not words:
                continue
            rows = _cluster_rows(words)

            slot = venue = ""
            for row in rows:
                text = " ".join(w["text"] for w in row)
                m = re.search(r"Date:\s*([A-Za-z0-9,/ ]+?)(?:\s{2,}|Slot|$)", text)
                if m:
                    date = m.group(1).strip()
                m = re.search(r"Slot:\s*([0-9:apmAPM.\- ]+?)(?:\s{2,}|Venue|$)", text)
                if m:
                    slot = m.group(1).strip()
                m = re.search(r"Venue:\s*(.+?)\s*$", text)
                if m:
                    venue = m.group(1).strip()

            split = None
            for row in rows:
                heads = [w for w in row if w["text"] == "S#"]
                if len(heads) >= 2:
                    split = heads[1]["x0"] - 5
                    break
            if split is None:
                split = float(page.width) / 2

            left_halves, right_halves = [], []
            for row in rows:
                left = " ".join(w["text"] for w in row if w["x0"] < split).strip()
                right = " ".join(w["text"] for w in row if w["x0"] >= split).strip()
                if left:
                    left_halves.append(left)
                if right:
                    right_halves.append(right)

            paper = ""
            for half in left_halves + right_halves:
                rec = FAST_REC.match(half)
                if rec:
                    students.append({
                        "name": re.sub(r"\s+", " ", rec.group(3).strip()),
                        "nuid": rec.group(2).upper(),
                        "seat": rec.group(4),
                        "paper": paper,
                        "time": slot,
                        "class": venue,
                    })
                    continue
                stripped = half.strip()
                paper_match = FAST_PAPER.match(half)
                if paper_match and re.match(r"^[A-Z]{2,3}\d{3,4}", stripped):
                    paper = _clean_paper(paper_match.group(1), paper_match.group(2))
                elif (paper and re.fullmatch(r"[A-Za-z][A-Za-z \-]*", stripped)
                      and not re.search(r"(Lab|Development)$", paper)):
                    paper = re.sub(r"\s+", " ", f"{paper} {stripped}")

    return students, date


def parse_fast_seating(text: str) -> tuple[list[dict[str, str]], str]:
    """Parse a FAST seating sheet. Returns (students, exam_date).

    The page flows column-major: the whole LEFT column is read top-to-bottom,
    then the whole RIGHT column. A section header (e.g. "CL1002,BCS-2C") can sit
    at the bottom of the left column and apply to the right column, so each venue
    block is processed as left-column lines followed by right-column lines, with
    the current paper/section carried across that boundary.
    """
    date = slot = ""
    blocks: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None

    for raw in text.splitlines():
        line = raw.rstrip()

        m = re.search(r"Date:\s*([A-Za-z0-9,/ ]+?)(?:\s{2,}|Slot|$)", line)
        if m:
            date = m.group(1).strip()
        m = re.search(r"Slot:\s*([0-9:apmAPM.\- ]+?)(?:\s{2,}|Venue|$)", line)
        if m:
            slot = m.group(1).strip()
        m = re.search(r"Venue:\s*(.+?)\s*$", line)
        if m:
            current = {"venue": m.group(1).strip(), "slot": slot, "lines": [], "split": 43}
            blocks.append(current)
            continue
        if current is None:
            continue
        if re.search(r"S#\s+Roll", line):
            second = line.find("S#", line.find("S#") + 1)
            if second > 0:
                current["split"] = second
            continue
        current["lines"].append(line)

    students: list[dict[str, str]] = []
    for block in blocks:
        split = block["split"]
        left = [ln[:split] for ln in block["lines"]]
        right = [ln[split:] for ln in block["lines"]]
        paper = ""
        for half in left + right:  # column-major: full left column, then full right
            rec = FAST_REC.match(half)
            if rec:
                name = re.sub(r"\s+\d{1,3}$", "", rec.group(3).strip())
                name = re.sub(r"\s+", " ", name).strip()
                students.append({
                    "name": name,
                    "nuid": rec.group(2).upper(),
                    "seat": rec.group(4),
                    "paper": paper,
                    "time": block["slot"],
                    "class": block["venue"],
                })
                continue
            stripped = half.strip()
            paper_match = FAST_PAPER.match(half)
            if paper_match and re.match(r"^[A-Z]{2,3}\d{3,4}", stripped):
                paper = _clean_paper(paper_match.group(1), paper_match.group(2))
            elif paper and re.fullmatch(r"[A-Za-z][A-Za-z \-]*", stripped) and not paper.endswith("Lab"):
                paper = re.sub(r"\s+", " ", f"{paper} {stripped}")

    return students, date


def parse_seating_plan_email(body: str, subject: str, pdf_bytes: bytes | None = None) -> dict[str, Any]:
    students: list[dict[str, str]] = []
    exam_date = ""

    # 1) Coordinate-based parse (most accurate: keeps seats aligned to rows).
    if pdf_bytes:
        try:
            students, exam_date = parse_pdf_coordinates(pdf_bytes)
        except Exception as exc:
            print(f"Coordinate parse failed ({exc}); falling back to text parse.")
            students = []

    # 2) Text-layout FAST parser, then 3) generic table / key-value.
    if not students:
        students, exam_date = parse_fast_seating(body)
    if not students:
        students = parse_tabular_body(body) or parse_key_value_body(body)

    cleaned = [normalize_entry(s) for s in students]
    cleaned = [s for s in cleaned if s.get("name") or s.get("nuid")]

    # De-duplicate (same student can appear once per parse path / page).
    seen: set[tuple] = set()
    deduped = []
    for s in cleaned:
        key = (s.get("nuid", ""), s.get("seat", ""), s.get("class", ""), s.get("name", ""))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(s)
    cleaned = deduped

    if not cleaned:
        raise ValueError(
            "Could not parse any student records from the email PDF payload. "
            "Verify table alignment formats inside the source file."
        )

    document: dict[str, Any] = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "source_subject": subject,
        "count": len(cleaned),
        "students": cleaned,
    }
    if exam_date:
        document["exam_date"] = exam_date
    return document


# ── FAST Final Exam Schedule parser (xlsx, matrix layout) ──────────────────────
# Each school sheet (FSC Final / FSM Final / FSE Final) is a matrix: column A
# holds the exam date (present once per date block, blank on subsequent rows of
# that block), and columns B/D/F hold the three time-slot cells (9-12, 1-4, 5-8),
# each stacked with one paper per non-blank row. A paper cell looks like:
#   "CS2005 Database Systems\nBS(CS) (A,B,C,D,E,F,G)\nBS(AI) (A,B,C)\n2024"
# with the batch year and any department/section group on their own lines.

EXAM_SLOT_COLUMNS = {2: "9:00 to 12:00 PM", 4: "1:00 to 4:00 PM", 6: "5:00 to 8:00 PM"}  # 1-indexed: B,D,F
EXAM_SCHOOL_SHEETS = [(re.compile(r"FSC", re.I), "computing"),
                      (re.compile(r"FSM", re.I), "business"),
                      (re.compile(r"FSE", re.I), "engineering")]

EXAM_CODE_RE = re.compile(r"^([A-Z]{2,3}\d{3,4})\s*(.*)$", re.S)
EXAM_YEAR_LINE_RE = re.compile(r"^(20\d{2})$")
EXAM_YEAR_ANY_RE = re.compile(r"(20\d{2})")
EXAM_DEPT_RE = re.compile(r"^BS\(([A-Z]+)\)\s*(.*)$", re.S)
EXAM_NOTE_RE = re.compile(r"\(in[^)]*\)", re.I)


def _excel_serial_to_iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value.isoformat()
    try:
        n = float(value)
    except (TypeError, ValueError):
        return ""
    from datetime import date, timedelta
    return (date(1899, 12, 30) + timedelta(days=round(n))).isoformat()


def _iso_date_day_name(iso_date: str) -> str:
    if not iso_date:
        return ""
    from datetime import date as _date
    y, m, d = (int(x) for x in iso_date.split("-"))
    return _date(y, m, d).strftime("%A")


def parse_exam_paper_cell(text: str) -> dict[str, Any] | None:
    """Parse one matrix cell into {code, course, batch, sections, notes}."""
    lines = [ln.strip() for ln in re.split(r"\r\n|\r|\n", text) if ln.strip()]
    if not lines:
        return None

    m = EXAM_CODE_RE.match(lines[0])
    code = m.group(1) if m else None
    course = m.group(2).strip() if m else lines[0]
    year: str | None = None
    notes: list[str] = []
    sections: dict[str, list[str]] = {}

    for ln in lines[1:]:
        if EXAM_YEAR_LINE_RE.match(ln):
            year = ln
            continue
        if "BS(" in ln:
            for chunk in re.split(r"(?=BS\()", ln):
                chunk = chunk.strip()
                if not chunk:
                    continue
                dm = EXAM_DEPT_RE.match(chunk)
                if not dm:
                    continue
                dept = dm.group(1)
                remainder = dm.group(2).strip()
                note_m = EXAM_NOTE_RE.search(remainder)
                if note_m:
                    notes.append(note_m.group(0))
                    remainder = remainder.replace(note_m.group(0), "").strip()
                remainder = re.sub(r"^-", "", remainder).strip()
                remainder = re.sub(r"^\(|\)$", "", remainder).strip()
                remainder = re.sub(r"\s+", "", remainder)
                if "," in remainder:
                    secs = [s for s in remainder.split(",") if s]
                elif re.fullmatch(r"[A-Z]{2,}", remainder or ""):
                    secs = list(remainder)
                elif remainder:
                    secs = [remainder]
                else:
                    secs = []
                secs = sorted({s.strip().upper() for s in secs if s.strip()})
                sections[dept] = sorted(set(sections.get(dept, [])) | set(secs))
        else:
            ym = EXAM_YEAR_ANY_RE.search(ln)
            if ym and not year:
                year = ym.group(1)
            else:
                notes.append(ln)

    return {
        "code": code,
        "course": course,
        "batch": year,
        "sections": sections,
        "notes": "; ".join(notes) if notes else None,
    }


def parse_exam_schedule_sheet(ws: Any) -> list[dict[str, Any]]:
    exams: list[dict[str, Any]] = []
    current_date = ""
    for row in ws.iter_rows():
        if len(row) > 0 and row[0].value is not None:
            iso = _excel_serial_to_iso_date(row[0].value)
            if iso:
                current_date = iso
        for col_idx, slot_label in EXAM_SLOT_COLUMNS.items():
            if col_idx - 1 >= len(row):
                continue
            cell_value = row[col_idx - 1].value
            if cell_value is None:
                continue
            parsed = parse_exam_paper_cell(str(cell_value))
            if not parsed or not parsed.get("code"):
                continue
            entry: dict[str, Any] = {
                "date": current_date or None,
                "day": _iso_date_day_name(current_date) or None,
                "time": slot_label,
                "code": parsed["code"],
                "course": parsed["course"],
                "batch": parsed["batch"],
                "sections": parsed["sections"],
            }
            if parsed.get("notes"):
                entry["notes"] = parsed["notes"]
            exams.append(entry)
    return exams


# ── FAST Midterm/Sessional Exam Schedule parser (room-grid layout) ─────────────
# Some "schedule" emails are a Midterm/Sessional schedule, not the Final Exam
# Schedule - a completely different layout: one row per ROOM (not per stacked
# paper), columns are Date (A) / Room (B) / one course per time-slot (C, E, G).
# Course cells have NO department/section info at all (e.g. just
# "MT-1003 Calculus and Analytical Geometry") - there's nothing to filter by
# dept/batch/section, so this is shown as a single flat table instead. A
# course spanning multiple rooms is a merged cell in the source (only the
# top-left cell of the merge carries the value), so merges must be forward-
# filled before reading the grid.

MIDTERM_HEADER_TIME_RE = re.compile(r"^\d{1,2}:\d{2}\s*(?:-|to)\s*\d{1,2}:\d{2}", re.I)
MIDTERM_CODE_RE = re.compile(r"^([A-Z]{2,3}-?\d{3,4})\s+(.*)$", re.S)
# A trailing full-row note (e.g. "Midterm Exams of ... will be held later.")
# is often a merge spanning ALL columns, which would otherwise forward-fill
# into the room/course columns too and look like fake data - require the
# room column to at least START with a room-code-like prefix (e.g. "C-301",
# "A-101  1st flr", "B-019") to guard against that.
MIDTERM_ROOM_RE = re.compile(r"^[A-Z]{1,3}-?\d+")
MIDTERM_SLOT_COLS = (2, 4, 6)  # 0-indexed: C, E, G


def _fill_merged_values(ws: Any) -> list[list[Any]]:
    """Materialize a worksheet into a 2D value grid, forward-filling merged
    ranges so every cell in a merge carries the top-left anchor's value
    (openpyxl only stores the value on that anchor cell)."""
    rows = list(ws.iter_rows())
    grid: list[list[Any]] = [[c.value for c in row] for row in rows]
    try:
        merges = list(ws.merged_cells.ranges)
    except Exception:
        merges = []
    for rng in merges:
        if rng.min_row - 1 >= len(grid) or rng.min_col - 1 >= len(grid[rng.min_row - 1]):
            continue
        anchor_val = grid[rng.min_row - 1][rng.min_col - 1]
        for r in range(rng.min_row, rng.max_row + 1):
            if r - 1 >= len(grid):
                continue
            row_list = grid[r - 1]
            for c in range(rng.min_col, rng.max_col + 1):
                c0 = c - 1
                while len(row_list) <= c0:
                    row_list.append(None)
                if row_list[c0] is None:
                    row_list[c0] = anchor_val
    return grid


def parse_midterm_schedule_sheet(ws: Any) -> list[dict[str, Any]]:
    grid = _fill_merged_values(ws)

    header_idx = next(
        (i for i, row in enumerate(grid) if len(row) > 2 and MIDTERM_HEADER_TIME_RE.match(str(row[2] or ""))),
        None,
    )
    if header_idx is None:
        return []
    slot_labels = {
        c: str(grid[header_idx][c]).strip()
        for c in MIDTERM_SLOT_COLS if c < len(grid[header_idx]) and grid[header_idx][c]
    }

    exams: list[dict[str, Any]] = []
    current_date = ""
    for row in grid[header_idx + 1:]:
        date_val = row[0] if len(row) > 0 else None
        if date_val is not None:
            iso = _excel_serial_to_iso_date(date_val)
            if iso:
                current_date = iso
        room = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if not room or not MIDTERM_ROOM_RE.match(room):
            continue
        for col in MIDTERM_SLOT_COLS:
            if col >= len(row) or not row[col]:
                continue
            cell_text = str(row[col]).strip()
            m = MIDTERM_CODE_RE.match(cell_text)
            exams.append({
                "date": current_date or None,
                "day": _iso_date_day_name(current_date) or None,
                "time": slot_labels.get(col, ""),
                "room": room,
                "code": m.group(1) if m else None,
                "course": m.group(2).strip() if m else cell_text,
            })
    return exams


def parse_exam_schedule_workbook(xlsx_bytes: bytes, subject: str, filename: str = "") -> dict[str, dict[str, Any]]:
    """Parse an exam-schedule workbook. Returns {school: document}.

    Tries the Final Exam Schedule's dept/section matrix format first; if a
    sheet yields nothing there (e.g. a Midterm/Sessional schedule, which uses
    a different room-grid layout with no dept/section info at all), falls
    back to the flat room-grid parser for that same sheet. The resulting
    document always has both "exams" (dept/batch-filterable) and
    "flat_exams" (shown as one plain table, no filtering) - whichever format
    was actually found populates one of the two; the other stays empty.
    """
    import openpyxl  # local import: only needed for the exam-schedule route

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    documents: dict[str, dict[str, Any]] = {}
    try:
        for sheet_name in wb.sheetnames:
            school = next((key for pattern, key in EXAM_SCHOOL_SHEETS if pattern.search(sheet_name)), None)
            if not school:
                continue
            ws = wb[sheet_name]
            exams = parse_exam_schedule_sheet(ws)
            flat_exams = [] if exams else parse_midterm_schedule_sheet(ws)
            if not exams and not flat_exams:
                continue
            documents[school] = {
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "source_subject": subject,
                "source_filename": filename,
                "count": len(exams) + len(flat_exams),
                "exams": exams,
                "flat_exams": flat_exams,
            }
    finally:
        wb.close()

    if not documents:
        raise ValueError(
            "Could not find a recognizable exam-schedule sheet (FSC/FSM/FSE) "
            "in the attached workbook, or no papers could be parsed from it."
        )
    return documents


# ── FAST Show Up Schedule parser (plain table layout) ───────────────────────────
# Unlike the Final Exam Schedule (a date/time-slot matrix), the Show Up Schedule
# workbook is a normal one-row-per-section table with columns:
#   Sr No. | Code | Course | Section | Teacher Name | Total Students |
#   Exam Date | Show-Up Date | Show-Up Time | Show-Up Venue
# The "Section" cell encodes program+semester+section+subsection, e.g.:
#   "BCS-3B"   -> Bachelor's, CS, semester 3, section B
#   "BCS-1J1"  -> Bachelor's, CS, semester 1, section J, subsection 1 (paired
#                 with a "BCS-1J2" row on a different date/time/venue — these
#                 are kept as separate exam entries, not merged, since a show-up
#                 slot is genuinely per-subsection).
# Only "B" (Bachelor's) rows are kept — "M"/"P" (Masters/PhD) rows are outside
# this app's BS-only department scope, same as the Final Exam Schedule feature.

SHOWUP_HEADER_RE = re.compile(r"^\s*code\s*$", re.I)
SHOWUP_SECTION_RE = re.compile(r"^-?([BMP]?)([A-Z]{2})-(\d+)([A-Z])(\d?)$")
SHOWUP_TERM_YEAR_RE = re.compile(r"(20\d{2})")


def _showup_semester_to_batch(semester: int, current_entry_year: int) -> str:
    """Map a semester number to an admission ("batch") year, e.g. for a
    Fall-2025 term: semester 1 -> 2025, semester 3 -> 2024, semester 5 -> 2023,
    semester 7 -> 2022 (each pair of semesters = one admission year)."""
    return str(current_entry_year - (semester - 1) // 2)


def parse_showup_schedule_sheet(ws: Any, current_entry_year: int) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows())
    header_row_idx = next(
        (i for i, row in enumerate(rows) if len(row) > 1 and SHOWUP_HEADER_RE.match(str(row[1].value or ""))),
        None,
    )
    if header_row_idx is None:
        return []

    exams: list[dict[str, Any]] = []
    for row in rows[header_row_idx + 1:]:
        get = lambda idx: row[idx].value if idx < len(row) else None  # noqa: E731
        code = get(1)
        course = get(2)
        section_raw = get(3)
        teacher = get(4)
        exam_date_val = get(6)
        showup_date_val = get(7)
        showup_time = get(8)
        showup_venue = get(9)
        if not code or not section_raw:
            continue

        m = SHOWUP_SECTION_RE.match(str(section_raw).strip())
        if not m or m.group(1) != "B":
            continue  # skip Masters/PhD/unrecognized rows - BS-only scope
        dept, semester, section_letter, subsection = m.group(2), int(m.group(3)), m.group(4), m.group(5)
        section = section_letter + subsection

        showup_date_iso = _excel_serial_to_iso_date(showup_date_val) if showup_date_val is not None else ""
        exam_date_iso = _excel_serial_to_iso_date(exam_date_val) if exam_date_val is not None else ""

        entry: dict[str, Any] = {
            "date": showup_date_iso or None,
            "day": _iso_date_day_name(showup_date_iso) or None,
            "time": str(showup_time).strip() if showup_time else "",
            "venue": str(showup_venue).strip() if showup_venue else "",
            "code": str(code).strip(),
            "course": str(course).strip() if course else "",
            "batch": _showup_semester_to_batch(semester, current_entry_year),
            "sections": {dept: [section]},
        }
        if teacher:
            entry["teacher"] = str(teacher).strip()
        if exam_date_iso:
            entry["exam_date"] = exam_date_iso
        exams.append(entry)
    return exams


def parse_showup_schedule_workbook(xlsx_bytes: bytes, subject: str, filename: str = "") -> dict[str, dict[str, Any]]:
    """Parse the Show Up Schedule workbook. Returns {"computing": document}.

    Only the workbook's VISIBLE sheet(s) are parsed (the file typically keeps
    prior terms' sheets around, hidden) so this keeps working as the sheet
    name changes each term (e.g. "Final Examination, Fall-2025" next becomes
    "..., Spring-2026") without needing a name pattern to match against.
    """
    import openpyxl  # local import: only needed for the showup-schedule route

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    all_exams: list[dict[str, Any]] = []
    try:
        year_match = SHOWUP_TERM_YEAR_RE.search(filename) or SHOWUP_TERM_YEAR_RE.search(subject)
        current_entry_year = int(year_match.group(1)) if year_match else datetime.now(timezone.utc).year

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if getattr(ws, "sheet_state", "visible") != "visible":
                continue
            sheet_year_match = SHOWUP_TERM_YEAR_RE.search(sheet_name)
            entry_year = int(sheet_year_match.group(1)) if sheet_year_match else current_entry_year
            all_exams.extend(parse_showup_schedule_sheet(ws, entry_year))
    finally:
        wb.close()

    if not all_exams:
        raise ValueError(
            "Could not parse any Show Up Schedule rows from the attached "
            "workbook's visible sheet(s)."
        )

    return {
        "computing": {
            "updated_at": datetime.now(timezone.utc).isoformat(),
            "source_subject": subject,
            "source_filename": filename,
            "count": len(all_exams),
            "exams": all_exams,
        }
    }


# ── Show Up Schedule: live Google Sheet polling ─────────────────────────────────
# The school doesn't resend a new email when the schedule changes - they just
# edit the same Google Sheet that was linked in the one email they already
# sent. So instead of relying on new emails, we extract that sheet's link ONCE
# (from the existing email's body, scanned automatically) and from then on poll
# the live sheet directly on a tight schedule (see
# .github/workflows/poll-showup-sheet.yml), independent of email entirely.

SHOWUP_SHEET_SOURCE_PATH = "dbfolder/showup-sheet-source.json"
GOOGLE_SHEET_LINK_RE = re.compile(r"https://docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)")


def extract_google_sheet_id(text: str) -> str | None:
    m = GOOGLE_SHEET_LINK_RE.search(text or "")
    return m.group(1) if m else None


def read_showup_sheet_id() -> str | None:
    try:
        with open(SHOWUP_SHEET_SOURCE_PATH, encoding="utf-8") as fh:
            return (json.load(fh) or {}).get("sheet_id") or None
    except (FileNotFoundError, json.JSONDecodeError):
        return None


def maybe_bootstrap_showup_sheet_source(kind: str, subject: str, body: str) -> None:
    """If this is a showup-schedule email and its body contains a Google Sheets
    link we haven't recorded yet, save it to disk so the poller can use it.
    Safe to call every run - a no-op once the link is already recorded (or if
    none is found at all)."""
    if kind != "showup_schedule":
        return
    sheet_id = extract_google_sheet_id(body)
    if not sheet_id or sheet_id == read_showup_sheet_id():
        return
    _write_json_file(SHOWUP_SHEET_SOURCE_PATH, {
        "sheet_id": sheet_id,
        "discovered_at": datetime.now(timezone.utc).isoformat(),
        "discovered_from_subject": subject,
    })
    print(f"Discovered/updated showup schedule Google Sheet id: {sheet_id!r}")


def fetch_showup_sheet_export_bytes(sheet_id: str) -> bytes:
    """Download the live sheet as .xlsx via Google's public export endpoint.
    Requires the sheet to be shared as "anyone with the link can view" (same
    requirement as the class timetable sync)."""
    import requests  # already a project dependency (requirements.txt)

    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    resp = requests.get(url, timeout=30)
    if resp.status_code != 200 or not resp.content.startswith(b"PK"):
        raise RuntimeError(
            f"Could not download the live showup sheet (HTTP {resp.status_code}). "
            "It may not be shared as 'Anyone with the link can view'."
        )
    return resp.content


# ── GitHub REST API commit ─────────────────────────────────────────────────────

def github_request(method: str, url: str, token: str, payload: dict | None = None) -> Any:
    data = None
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "User-Agent": "vtable-seating-sync",
    }
    if payload is not None:
        data = json.dumps(payload).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = Request(url, data=data, headers=headers, method=method)
    try:
        with urlopen(req, timeout=30) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw) if raw else {}
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"GitHub API {exc.code}: {detail}") from exc
    except URLError as exc:
        raise RuntimeError(f"GitHub API network error: {exc}") from exc


def resolve_repo() -> tuple[str, str]:
    repo = os.environ.get("GH_REPO", "").strip()
    if repo and "/" in repo:
        owner, name = repo.split("/", 1)
        return owner, name
    owner = os.environ.get("VERCEL_GIT_REPO_OWNER", "").strip()
    slug = os.environ.get("VERCEL_GIT_REPO_SLUG", "").strip()
    if owner and slug:
        return owner, slug
    raise RuntimeError(
        "Set GH_REPO (owner/repo) or deploy via Vercel properties"
    )


def commit_json_to_github(document: dict[str, Any], repo_path: str) -> dict[str, Any]:
    """Overwrite `repo_path` completely with `document` via the GitHub contents API."""
    token = os.environ.get("GH_TOKEN", "").strip()
    if not token:
        raise RuntimeError("GH_TOKEN environment variable is required")

    owner, repo = resolve_repo()
    branch = os.environ.get("GH_BRANCH", "main").strip() or "main"
    api_base = f"https://api.github.com/repos/{owner}/{repo}/contents/{repo_path}"

    existing_sha: str | None = None
    try:
        meta = github_request("GET", f"{api_base}?ref={branch}", token)
        existing_sha = meta.get("sha")
    except RuntimeError as exc:
        if "404" not in str(exc):
            raise

    content_bytes = json.dumps(document, indent=2, ensure_ascii=False).encode("utf-8")
    payload: dict[str, Any] = {
        "message": f"Sync {repo_path} from Gmail ({document.get('count', 0)} entries)",
        "content": base64.b64encode(content_bytes).decode("ascii"),
        "branch": branch,
    }
    if existing_sha:
        payload["sha"] = existing_sha

    result = github_request("PUT", api_base, token, payload)
    return {
        "committed": True,
        "path": repo_path,
        "branch": branch,
        "sha": result.get("content", {}).get("sha"),
        "html_url": result.get("content", {}).get("html_url"),
    }


def _write_json_file(path: str, document: dict[str, Any]) -> None:
    directory = os.path.dirname(path)
    if directory:
        os.makedirs(directory, exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(document, fh, indent=2, ensure_ascii=False)


# ── Vercel handler ───────────────────────────────────────────────────────────

class handler(BaseHTTPRequestHandler):
    """Vercel Python entrypoint – trigger via GET or POST /api/fetch-timetable."""

    def do_GET(self) -> None:
        self._run_sync()

    def do_POST(self) -> None:
        self._run_sync()

    def _run_sync(self) -> None:
        try:
            kind, subject, body, attachment, attachment_filename = fetch_latest_matching_email()
            if kind == "seating":
                document = parse_seating_plan_email(body, subject, attachment)
                path = SUBJECT_ROUTES["seating"][1]
                commit_info = commit_json_to_github(document, path)
                json_response(
                    self, 200,
                    {
                        "ok": True,
                        "message": "Seating plan synced successfully from PDF file",
                        "students_parsed": document["count"],
                        "source_subject": subject,
                        "github": commit_info,
                    },
                )
            else:
                if not attachment:
                    raise RuntimeError(f"No .xlsx attachment found on the {kind} email")
                parser = (
                    parse_showup_schedule_workbook if kind == "showup_schedule"
                    else parse_exam_schedule_workbook
                )
                documents = parser(attachment, subject, attachment_filename)
                prefix = SCHEDULE_FILE_PREFIX[kind]
                github_results = {}
                for school, doc in documents.items():
                    path = f"dbfolder/{prefix}-{school}.json"
                    github_results[school] = commit_json_to_github(doc, path)
                json_response(
                    self, 200,
                    {
                        "ok": True,
                        "message": f"{kind} synced successfully from xlsx file",
                        "schools_parsed": {s: d["count"] for s, d in documents.items()},
                        "source_subject": subject,
                        "github": github_results,
                    },
                )
        except Exception as exc:
            json_response(self, 500, {"ok": False, "error": str(exc)})

    def log_message(self, format: str, *args: Any) -> None:
        return


# ── Standalone CLI (GitHub Actions / local) ────────────────────────────────────
# Same fetch+parse pipeline as the Vercel handler, but writes JSON straight to
# disk so a CI workflow can commit it with its own GITHUB_TOKEN. Only
# GMAIL_USER / GMAIL_PASS are required in this mode.

def run_cli() -> int:
    try:
        kind, subject, body, attachment, attachment_filename = fetch_latest_matching_email()
    except RuntimeError as exc:
        # A scheduled run with no matching email is a no-op, not a failure.
        if "No emails found" in str(exc):
            print("No seating/schedule/showup email found - nothing to sync.")
            return 0
        raise

    # Regardless of what else happens below: if this is a showup-schedule email
    # and its body links a Google Sheet we haven't recorded yet, save it so the
    # live poller (run_showup_poll_cli, on its own 5-min schedule) can use it -
    # this is how we track edits made directly in the sheet, with no new email.
    maybe_bootstrap_showup_sheet_source(kind, subject, body)

    if kind == "seating":
        document = parse_seating_plan_email(body, subject, attachment)
        path = SUBJECT_ROUTES["seating"][1]
        _write_json_file(path, document)
        print(f"Wrote {path}: {document['count']} student(s) (subject: {subject!r})")
    else:
        if not attachment:
            print(f"{kind} email (subject: {subject!r}) had no .xlsx attachment - skipping "
                  f"the attachment parse (any linked Google Sheet was still checked above).")
            return 0
        parser = (
            parse_showup_schedule_workbook if kind == "showup_schedule"
            else parse_exam_schedule_workbook
        )
        documents = parser(attachment, subject, attachment_filename)
        prefix = SCHEDULE_FILE_PREFIX[kind]
        for school, doc in documents.items():
            path = f"dbfolder/{prefix}-{school}.json"
            # Each new schedule email fully REPLACES the previous one - the
            # latest emailed schedule is the current one. (A dept/section
            # schedule like the Final/Sessional populates "exams"; a room-grid
            # Midterm populates "flat_exams"; the unused array stays empty so
            # the old format's data doesn't linger.)
            _write_json_file(path, doc)
            print(f"Wrote {path}: {doc['count']} exam entries (subject: {subject!r}, file: {attachment_filename!r})")
    return 0


# ── Standalone CLI: live Show Up Schedule sheet poll ────────────────────────────
# Runs on its own tight schedule (every 5 min - see
# .github/workflows/poll-showup-sheet.yml), independent of email. No-op if no
# sheet has been discovered yet (see maybe_bootstrap_showup_sheet_source above).

def run_showup_poll_cli() -> int:
    sheet_id = read_showup_sheet_id()
    if not sheet_id:
        print(f"No showup Google Sheet registered yet in {SHOWUP_SHEET_SOURCE_PATH} - nothing to poll.")
        return 0

    try:
        xlsx_bytes = fetch_showup_sheet_export_bytes(sheet_id)
    except RuntimeError as exc:
        # A transient fetch failure shouldn't fail the whole scheduled run.
        print(f"Could not fetch live showup sheet {sheet_id!r}: {exc}")
        return 0

    documents = parse_showup_schedule_workbook(
        xlsx_bytes, subject="Live Google Sheet poll", filename=f"Google Sheet ({sheet_id})",
    )
    for school, doc in documents.items():
        path = f"dbfolder/showup-schedule-{school}.json"
        _write_json_file(path, doc)
        print(f"Wrote {path}: {doc['count']} exam entries (polled live sheet {sheet_id!r})")
    return 0


if __name__ == "__main__":
    import sys
    if "--poll-showup" in sys.argv:
        sys.exit(run_showup_poll_cli())
    sys.exit(run_cli())